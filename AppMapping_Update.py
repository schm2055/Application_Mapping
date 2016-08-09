from openpyxl import Workbook, load_workbook
import os, ctypes, getpass, easygui
import sqlite3

#Connect to SQL DB - creates DB in the directory if it does not exist at the time of execution
conn_db = sqlite3.connect('appmapping.sqlite')
cur_db = conn_db.cursor()
print 'Database created and/or connected to successfully!'

#Create additional SQL table for applications and events - 1 event per 1 application
cur_db.executescript('''

CREATE TABLE IF NOT EXISTS App_Events(
	Application TEXT,
	Event TEXT,
);

''')


#Load Applications mapped to events worksheet
while True:
	ctypes.windll.user32.MessageBoxA(0, 'Open your Events Document', 'Open', 1)
	events = easygui.fileopenbox()
	try:
		events_wb = load_workbook(filename = events, use_iterators = True)
		break
	except:
		ctypes.windll.user32.MessageBoxA(0, 'Please enter a valid events inventory  name', 'Import Error!', 1)
		continue
events_active_sheet = events_wb.get_sheet_names()[0]
events_working_sheet = events_wb.get_sheet_by_name(events_active_sheet) 

#Define function to get row values in a list
def iter_rows(working_sheet):
    for row in working_sheet.iter_rows():
        yield [cell.value for cell in row]
		
#Import Events Contents into App_Events Table		
services_row_list = iter_rows(events_working_sheet)

for list in services_row_list:
	application = list[0]
	event = list[1]
	cur_db.execute('''INSERT INTO App_Events (Application, Event)
		VALUES (?, ?)''', (application, event, ))
	conn_db.commit()

print 'Services Imported Successfully!'


#Add events to applications in service_instances
cur_db.execute('''SELECT DISTINCT Application_Services FROM Service_Instances''')
app_list = [record[0] for record in cur_db.fetchall()]

for app in app_list:
	cur_db.execute("SELECT Event FROM App_Events WHERE Application = ?", (app, ))
	e_list = [record[0] for record in cur_db.fetchall()]
	try:
		e_val = e_list[0]
	except:
		continue
	cur_db.execute("UPDATE Service_Instances SET Event = ? WHERE Application = ?", (e_val, app, ))
	conn_db.commit()


#Success message
ctypes.windll.user32.MessageBoxA(0, 'Events Mapped Successfully!', 'Success!', 1)
